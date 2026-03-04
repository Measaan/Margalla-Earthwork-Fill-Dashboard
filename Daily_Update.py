#!/usr/bin/env python3
"""
Margalla Enclave Link Road - Daily Fill Updater
Enter 1 quantity per contractor per day. That's it.
"""
import os, sys, re
from datetime import date, datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
except ImportError:
    os.system("pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(SCRIPT_DIR, "Road_Data.xlsx")
DASH_FILE   = os.path.join(SCRIPT_DIR, "Margalla_Fill_Dashboard.html")
TARGET_DATE = "2026-03-18"

GROSS      = 898453
STRUCT_DED = 57294
NET_TARGET = GROSS - STRUCT_DED   # 841,159

CONTRACTORS = [
    ('Al Ajal Builders',  'Section-1 (0+000 to 1+000)', 242029),
    ('BAS Construction',  'Section-2 (1+000 to 2+000)', 349638),
    ('M. Shehzad Abbasi', 'Section-3 (2+000 to 3+000)', 211523),
    ('Fazal Rehman',      'Section-4 (3+000 to 3+750)',  95263),
]
C_COLORS = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444']

def clr(code,t): return "\033["+code+"m"+str(t)+"\033[0m"
def green(t):  return clr("32;1",t)
def yellow(t): return clr("33;1",t)
def cyan(t):   return clr("36;1",t)
def red(t):    return clr("31;1",t)
def bold(t):   return clr("1",t)
def fmt(n):    return "{:,}".format(int(round(n)))

# ── Load current WD from Excel ────────────────────────
def load_data():
    if not os.path.exists(EXCEL_FILE):
        print(red("\n  ERROR: Road_Data.xlsx not found!")); input("\n  Press Enter..."); sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Progress_Data"]
    data = []
    for row in ws.iter_rows(min_row=3, max_row=6, values_only=True):
        wd = float(row[3] or 0)   # col D = Work Done To Date
        data.append(wd)
    return wb, data

# ── Save updated WD to Excel ──────────────────────────
def save_data(wb, wd_list, daily_list, today_str):
    ws = wb["Progress_Data"]
    thin = Side(style='thin', color="AAAAAA")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    yf   = PatternFill("solid", fgColor="FFFDE7")
    for i, (wd, da) in enumerate(zip(wd_list, daily_list)):
        row = i + 3
        ws.cell(row=row, column=4).value = wd   # Work Done To Date
        ws.cell(row=row, column=5).value = da   # Daily Added

    ws_log   = wb["Daily_Log"]
    next_row = max(ws_log.max_row + 1, 3)
    for i, (cname, sec, qty) in enumerate(CONTRACTORS):
        if daily_list[i] > 0:
            for ci, v in enumerate([today_str, cname, sec, daily_list[i], wd_list[i], ""], 1):
                c = ws_log.cell(row=next_row, column=ci, value=v)
                c.border = bdr; c.fill = yf
                c.alignment = Alignment(horizontal='center', vertical='center')
                if ci in (4, 5): c.number_format = '#,##0'
            next_row += 1
    wb.save(EXCEL_FILE)

# ── Update dashboard HTML ─────────────────────────────
def update_dashboard(wd_list, daily_list):
    if not os.path.exists(DASH_FILE):
        print(red("  ERROR: Margalla_Fill_Dashboard.html not found!")); return

    now_str = datetime.now().strftime("%d %b %Y, %I:%M %p")

    lines = []
    for i, (cname, sec, qty) in enumerate(CONTRACTORS):
        wd = round(wd_list[i], 2)
        da = round(daily_list[i], 2)
        lines.append(
            '  {c:"' + cname + '",sec:"' + sec + '",qty:' + str(qty) +
            ',wd:' + str(wd) + ',da:' + str(da) + '}'
        )
    new_raw = 'const RAW=[\n' + ',\n'.join(lines) + '\n];'

    html = open(DASH_FILE, encoding='utf-8').read()
    html = re.sub(r'const RAW=\[.*?\];', new_raw, html, flags=re.DOTALL)
    html = re.sub(r'Updated: [^<]+</div>', 'Updated: ' + now_str + '</div>', html)
    open(DASH_FILE, 'w', encoding='utf-8').write(html)

# ── Main ──────────────────────────────────────────────
def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    print(bold("=" * 60))
    print(bold("   MARGALLA ENCLAVE LINK ROAD — DAILY FILL UPDATER"))
    print(bold("=" * 60))
    print(f"  Gross Qty    : {yellow(fmt(GROSS))} m\u00b3")
    print(f"  Struct Ded   : {red(fmt(STRUCT_DED))} m\u00b3")
    print(f"  Net Target   : {green(fmt(NET_TARGET))} m\u00b3")
    print(bold("=" * 60)); print()

    wb, wd_list = load_data()
    total_wd  = sum(wd_list)
    net_bal   = NET_TARGET - total_wd
    net_pct   = total_wd / NET_TARGET * 100

    print(f"  Work Done    : {green(fmt(total_wd))} m\u00b3")
    print(f"  Balance      : {red(fmt(net_bal))} m\u00b3")
    print(f"  Progress     : {yellow(f'{net_pct:.1f}%')}")
    print()
    print("  Per Contractor:")
    for i, (cname, sec, qty) in enumerate(CONTRACTORS):
        wd  = wd_list[i]
        pct = wd / qty * 100 if qty else 0
        bar = green if pct >= 70 else (yellow if pct >= 40 else red)
        print(f"    {cname:<24}  WD: {bar(fmt(wd)):>12} m\u00b3  ({pct:.1f}%)")
    print()

    today_str = input(cyan("  Enter date (YYYY-MM-DD) or press Enter for today: ")).strip()
    if not today_str: today_str = str(date.today())
    print()

    daily_list = []
    total_today = 0.0
    print(bold("  Enter today's quantity for each contractor:"))
    print(bold("  (Press Enter to skip / enter 0 for no work done)"))
    print()
    for i, (cname, sec, qty) in enumerate(CONTRACTORS):
        wd  = wd_list[i]
        bal = max(0, qty - wd)
        pct = wd / qty * 100 if qty else 0
        label = f"    {cname:<24}  Done:{fmt(wd):>10}  Bal:{fmt(bal):>10}  {pct:5.1f}%  Today: "
        val = input(yellow(label)).strip()
        try:    v = float(val) if val else 0.0
        except: v = 0.0
        daily_list.append(v)
        total_today += v
    print()

    # Update wd_list with today's additions
    new_wd = [wd_list[i] + daily_list[i] for i in range(4)]
    new_total = sum(new_wd)

    print(bold("=" * 60))
    print(f"  Today Added  : {green(fmt(total_today))} m\u00b3")
    print(f"  New Total WD : {green(fmt(new_total))} m\u00b3  ({yellow(f'{new_total/NET_TARGET*100:.1f}%')})")
    print(f"  New Balance  : {red(fmt(NET_TARGET - new_total))} m\u00b3")
    print(bold("=" * 60)); print()

    confirm = input(cyan("  Save and update dashboard? (Y/N): ")).strip().upper()
    if confirm != 'Y':
        print(yellow("\n  Cancelled. Nothing saved.")); input("\n  Press Enter..."); return

    save_data(wb, new_wd, daily_list, today_str)
    update_dashboard(new_wd, daily_list)

    print(green(f"\n  Road_Data.xlsx updated!"))
    print(green(f"  Margalla_Fill_Dashboard.html updated!"))
    print(f"\n  Open {bold('Margalla_Fill_Dashboard.html')} in browser to view.\n")
    input("  Press Enter to exit...")

if __name__ == '__main__':
    main()
